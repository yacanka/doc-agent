from pathlib import Path

import pytest
from docx import Document

from app.agent.executor import Executor
from app.agent.planner import Planner


class DummyLlm:
    def __init__(self, response):
        self.response = response

    def plan(self, instruction: str, document_text: str):
        return self.response


def _docx(path: Path, text: str) -> None:
    document = Document()
    document.add_paragraph(text)
    document.save(path)


def test_planner_accepts_valid_replacement_plan() -> None:
    planner = Planner(DummyLlm('{"operations":[{"action":"replace_text","tool":"word.replace_text","parameters":{"replacements":{"Alice":"Bob"}}}]}'))

    plan = planner.plan_operations("Replace Alice with Bob", "Hello Alice")

    assert plan.operations[0].parameters["replacements"] == {"Alice": "Bob"}


def test_planner_rejects_malformed_json() -> None:
    planner = Planner(DummyLlm('{"operations":'))

    with pytest.raises(ValueError, match="valid JSON"):
        planner.plan_operations("Replace Alice with Bob", "Hello Alice")


def test_planner_extracts_json_from_model_explanation() -> None:
    response = 'Sure {not json} then {"replacements":{"Alice":"Bob"}} done'
    planner = Planner(DummyLlm(response))

    replacements = planner.plan("Replace Alice with Bob", "Hello Alice")

    assert replacements == {"Alice": "Bob"}


def test_planner_rejects_unknown_tools() -> None:
    planner = Planner(DummyLlm({"operations": [{"action": "replace_text", "tool": "shell", "parameters": {"replacements": {"A": "B"}}}]}))

    with pytest.raises(ValueError, match="Invalid operation plan"):
        planner.plan_operations("Replace A with B", "A")


def test_planner_rejects_unsupported_actions() -> None:
    planner = Planner(DummyLlm({"operations": [{"action": "delete_file", "tool": "word.replace_text", "parameters": {}}]}))

    with pytest.raises(ValueError, match="Invalid operation plan"):
        planner.plan_operations("Delete the file", "A")


def test_planner_rejects_vague_destructive_instruction() -> None:
    planner = Planner(DummyLlm({"operations": []}))

    with pytest.raises(ValueError, match="Vague destructive"):
        planner.plan_operations("Delete everything", "Hello Alice")


def test_executor_uses_controlled_paths_for_valid_operations(tmp_path: Path) -> None:
    source = tmp_path / "source.docx"
    _docx(source, "Hello Alice")
    session = {"original_path": str(source), "id": "session", "filename": "source.docx"}
    planner = Planner(DummyLlm({"replacements": {"Alice": "Bob"}}))

    result = Executor(tmp_path / "outputs").apply_plan(session, planner.plan_operations("Replace Alice with Bob", "Hello Alice"))

    assert result.changed_count == 1
    assert result.report["valid"] is True
    assert Path(result.output_path).is_file()
    assert str(tmp_path / "outputs") in result.output_path


def _xlsx(path: Path, text: str) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet["A1"] = text
    sheet["B1"] = "=SUM(1,2)"
    workbook.save(path)


def test_executor_replaces_text_in_xlsx_without_touching_formulas(tmp_path: Path) -> None:
    from openpyxl import load_workbook

    source = tmp_path / "source.xlsx"
    _xlsx(source, "Hello Alice")
    session = {"original_path": str(source), "id": "session", "filename": "source.xlsx"}
    planner = Planner(DummyLlm({"replacements": {"Alice": "Bob"}}))

    result = Executor(tmp_path / "outputs").apply_plan(session, planner.plan_operations("Replace Alice", "Hello Alice"))

    workbook = load_workbook(result.output_path, data_only=False)
    assert result.changed_count == 1
    assert result.report["valid"] is True
    assert workbook["Data"]["A1"].value == "Hello Bob"
    assert workbook["Data"]["B1"].value == "=SUM(1,2)"


def test_planner_accepts_excel_update_cells_plan() -> None:
    planner = Planner(DummyLlm({
        "operations": [{
            "action": "update_cells",
            "tool": "excel.update_cells",
            "parameters": {"updates": {"Data": {"A2": "Bob"}}},
        }]
    }))

    plan = planner.plan_operations("Set Data A2 to Bob", "[Data]\nAlice")

    assert plan.operations[0].parameters["updates"] == {"Data": {"A2": "Bob"}}
