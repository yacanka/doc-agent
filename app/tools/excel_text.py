"""Text extraction and replacement helpers for XLSX workbooks."""
from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

_MAX_CONTEXT_CELLS = 500


def extract_text(path: Path) -> str:
    """Extract workbook context with sheet names, coordinates, values, and formulas."""
    workbook = load_workbook(path, data_only=False, read_only=True)
    try:
        return "\n".join(_sheet_text(sheet) for sheet in workbook.worksheets)
    finally:
        workbook.close()


def replace_workbook_text(workbook: Any, replacements: dict[str, str]) -> int:
    """Replace literal text in string cells across all worksheets."""
    changed_count = 0
    for sheet in workbook.worksheets:
        changed_count += _replace_sheet_text(sheet, replacements)
    return changed_count


def _sheet_text(sheet: Worksheet) -> str:
    lines = [_sheet_header(sheet)]
    lines.extend(_cell_lines(sheet))
    if len(lines) == 1:
        lines.append("<empty sheet>")
    return "\n".join(lines)


def _sheet_header(sheet: Worksheet) -> str:
    return f"Sheet '{sheet.title}' ({sheet.max_row} rows x {sheet.max_column} columns)"


def _cell_lines(sheet: Worksheet) -> list[str]:
    lines: list[str] = []
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                lines.append(_cell_line(cell))
            if len(lines) >= _MAX_CONTEXT_CELLS:
                return lines + ["... truncated workbook context ..."]
    return lines


def _cell_line(cell: Any) -> str:
    value = str(cell.value)
    value_type = "formula" if value.startswith("=") else "value"
    return f"{cell.parent.title}!{cell.coordinate} [{value_type}]: {value}"


def _replace_sheet_text(sheet: Worksheet, replacements: dict[str, str]) -> int:
    changed_count = 0
    for row in sheet.iter_rows():
        for cell in row:
            changed_count += _replace_cell_text(cell, replacements)
    return changed_count


def _replace_cell_text(cell: Any, replacements: dict[str, str]) -> int:
    if not isinstance(cell.value, str) or cell.value.startswith("="):
        return 0
    updated = _replacement_value(cell.value, replacements)
    if updated == cell.value:
        return 0
    cell.value = updated
    return 1


def _replacement_value(value: str, replacements: dict[str, str]) -> str:
    updated = value
    for source, target in replacements.items():
        updated = updated.replace(str(source), str(target))
    return updated
