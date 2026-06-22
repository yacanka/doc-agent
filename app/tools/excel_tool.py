"""Safe XLSX inspection and editing utilities."""
from __future__ import annotations

import copy
import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.tools.excel_text import extract_text, replace_workbook_text


@dataclass(frozen=True)
class SheetSummary:
    """Compact description of an Excel worksheet."""

    title: str
    max_row: int
    max_column: int
    merged_ranges: tuple[str, ...]
    freeze_panes: str | None
    auto_filter: str | None


@dataclass(frozen=True)
class WorkbookSummary:
    """Compact description of an Excel workbook."""

    sheets: tuple[SheetSummary, ...]
    creator: str | None
    modified_by: str | None
    title: str | None
    subject: str | None


def assert_xlsx(filename: str, content_type: str | None) -> None:
    """Validate that an upload is an XLSX workbook with an allowed content type."""
    allowed_types = {
        None,
        "",
        "application/octet-stream",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
    if not filename.lower().endswith(".xlsx") or content_type not in allowed_types:
        raise ValueError("Only XLSX workbook uploads are supported for Excel files")


def replace_text(source: Path, destination: Path, replacements: dict[str, str]) -> int:
    """Copy an XLSX and replace literal text inside string cells only."""
    shutil.copy2(source, destination)
    workbook = load_workbook(destination, data_only=False)
    changed_count = replace_workbook_text(workbook, replacements)
    workbook.save(destination)
    return changed_count


def inspect_workbook(path: Path) -> WorkbookSummary:
    """Read workbook structure without changing formulas, metadata, or layout."""
    workbook = load_workbook(path, data_only=False)
    sheets = tuple(_summarize_sheet(sheet) for sheet in workbook.worksheets)
    return WorkbookSummary(
        sheets=sheets,
        creator=_value(workbook.properties.creator),
        modified_by=_value(workbook.properties.lastModifiedBy),
        title=_value(workbook.properties.title),
        subject=_value(workbook.properties.subject),
    )


def extract_sheet_summaries(path: Path, sample_rows: int = 5) -> dict[str, dict[str, Any]]:
    """Return sheet dimensions, layout metadata, and a small value sample."""
    workbook = load_workbook(path, data_only=False)
    return {sheet.title: _sheet_payload(sheet, sample_rows) for sheet in workbook.worksheets}


def update_cells(source: Path, destination: Path, updates: dict[str, dict[str, Any]]) -> int:
    """Copy an XLSX and update specific cell values by sheet and coordinate."""
    shutil.copy2(source, destination)
    workbook = load_workbook(destination, data_only=False)
    changed_count = _apply_cell_updates(workbook, updates)
    workbook.save(destination)
    return changed_count


def append_rows(source: Path, destination: Path, sheet_name: str, rows: list[list[Any]]) -> int:
    """Copy an XLSX and append rows while copying styles from the row above."""
    shutil.copy2(source, destination)
    workbook = load_workbook(destination, data_only=False)
    sheet = _get_sheet(workbook, sheet_name)
    start_row = sheet.max_row + 1
    for offset, row_values in enumerate(rows):
        _append_row_with_style(sheet, start_row + offset, row_values)
    workbook.save(destination)
    return len(rows)


def copy_cell_style(source: Path, destination: Path, sheet_name: str, source_cell: str, target_cell: str) -> None:
    """Copy style attributes from a reference cell into another cell."""
    shutil.copy2(source, destination)
    workbook = load_workbook(destination, data_only=False)
    sheet = _get_sheet(workbook, sheet_name)
    _copy_style(sheet[source_cell], sheet[target_cell])
    workbook.save(destination)


def _summarize_sheet(sheet: Worksheet) -> SheetSummary:
    return SheetSummary(
        title=sheet.title,
        max_row=sheet.max_row,
        max_column=sheet.max_column,
        merged_ranges=tuple(str(rng) for rng in sheet.merged_cells.ranges),
        freeze_panes=_coordinate(sheet.freeze_panes),
        auto_filter=_value(sheet.auto_filter.ref),
    )


def _sheet_payload(sheet: Worksheet, sample_rows: int) -> dict[str, Any]:
    summary = _summarize_sheet(sheet)
    return {"summary": summary, "sample": _sample_values(sheet, sample_rows)}


def _sample_values(sheet: Worksheet, sample_rows: int) -> list[list[Any]]:
    limit = min(sheet.max_row, max(sample_rows, 0))
    return [[cell.value for cell in row] for row in sheet.iter_rows(max_row=limit)]


def _apply_cell_updates(workbook: Any, updates: dict[str, dict[str, Any]]) -> int:
    changed_count = 0
    for sheet_name, cells in updates.items():
        sheet = _get_sheet(workbook, sheet_name)
        for coordinate, value in cells.items():
            sheet[coordinate].value = value
            changed_count += 1
    return changed_count


def _append_row_with_style(sheet: Worksheet, row_number: int, row_values: list[Any]) -> None:
    _copy_row_layout(sheet, row_number)
    for column_index, value in enumerate(row_values, start=1):
        cell = sheet.cell(row=row_number, column=column_index, value=value)
        _copy_neighbor_style(sheet, cell)


def _copy_row_layout(sheet: Worksheet, row_number: int) -> None:
    if row_number <= 1:
        return
    source_height = sheet.row_dimensions[row_number - 1].height
    sheet.row_dimensions[row_number].height = source_height


def _copy_neighbor_style(sheet: Worksheet, cell: Any) -> None:
    if cell.row <= 1:
        return
    reference = sheet.cell(row=cell.row - 1, column=cell.column)
    _copy_style(reference, cell)


def _copy_style(source_cell: Any, target_cell: Any) -> None:
    target_cell._style = copy.copy(source_cell._style)
    target_cell.number_format = source_cell.number_format
    target_cell.protection = copy.copy(source_cell.protection)
    target_cell.alignment = copy.copy(source_cell.alignment)


def _get_sheet(workbook: Any, sheet_name: str) -> Worksheet:
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Worksheet not found: {sheet_name}")
    return workbook[sheet_name]


def _coordinate(value: Any) -> str | None:
    return getattr(value, "coordinate", value) if value else None


def _value(value: Any) -> str | None:
    return str(value) if value else None
